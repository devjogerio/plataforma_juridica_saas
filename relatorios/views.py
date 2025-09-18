from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, Http404
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Sum, Avg, Max, Min
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import date, datetime, timedelta
from decimal import Decimal
import json
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import TemplateRelatorio, ExecucaoRelatorio, DashboardPersonalizado, FiltroSalvo
from .forms import TemplateRelatorioForm, FiltroRelatorioForm, DashboardPersonalizadoForm, FiltroSalvoForm, ExportarRelatorioForm
from processos.models import Processo, Andamento, Prazo
from clientes.models import Cliente
from financeiro.models import Honorario, Despesa
from documentos.models import Documento
from usuarios.models import Usuario


class RelatoriosDashboardView(LoginRequiredMixin, TemplateView):
    """
    Dashboard principal do módulo de relatórios com templates e execuções recentes
    """
    template_name = 'relatorios/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Templates disponíveis
        if self.request.user.is_staff:
            templates = TemplateRelatorio.objects.filter(ativo=True)
        else:
            templates = TemplateRelatorio.objects.filter(
                Q(usuario_criador=self.request.user) | Q(publico=True),
                ativo=True
            )
        
        # Execuções recentes
        execucoes_recentes = ExecucaoRelatorio.objects.filter(
            usuario=self.request.user
        ).select_related('template').order_by('-data_execucao')[:10]
        
        # Estatísticas
        total_templates = templates.count()
        templates_publicos = templates.filter(publico=True).count()
        execucoes_hoje = ExecucaoRelatorio.objects.filter(
            usuario=self.request.user,
            data_execucao__date=date.today()
        ).count()
        
        # Templates mais utilizados
        templates_populares = templates.annotate(
            total_execucoes=Count('execucoes')
        ).filter(total_execucoes__gt=0).order_by('-total_execucoes')[:5]
        
        # Dashboards personalizados
        dashboards = DashboardPersonalizado.objects.filter(
            Q(usuario=self.request.user) | Q(publico=True),
            ativo=True
        ).order_by('-padrao', 'nome')[:5]
        
        context.update({
            'templates': templates[:8],
            'execucoes_recentes': execucoes_recentes,
            'total_templates': total_templates,
            'templates_publicos': templates_publicos,
            'execucoes_hoje': execucoes_hoje,
            'templates_populares': templates_populares,
            'dashboards': dashboards,
        })
        
        return context


class RelatorioProcessosView(LoginRequiredMixin, TemplateView):
    """
    Relatório específico de processos com filtros e análises
    """
    template_name = 'relatorios/processos.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aplicar filtros
        form = FiltroRelatorioForm(self.request.GET, user=self.request.user, tipo_relatorio='processos')
        context['form'] = form
        
        # Base queryset
        if self.request.user.is_staff:
            processos_qs = Processo.objects.all()
        else:
            processos_qs = Processo.objects.filter(responsavel=self.request.user)
        
        if form.is_valid():
            # Aplicar filtros do formulário
            processos_qs = self._aplicar_filtros(processos_qs, form.cleaned_data)
        
        # Estatísticas
        total_processos = processos_qs.count()
        processos_ativos = processos_qs.filter(status='ativo').count()
        processos_encerrados = processos_qs.filter(status='encerrado').count()
        valor_total_causas = processos_qs.aggregate(Sum('valor_causa'))['valor_causa__sum'] or Decimal('0.00')
        
        # Processos por status
        processos_por_status = list(
            processos_qs.values('status')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        # Processos por área do direito
        processos_por_area = list(
            processos_qs.values('area_direito__nome')
            .annotate(total=Count('id'))
            .order_by('-total')[:10]
        )
        
        # Processos por responsável
        processos_por_responsavel = list(
            processos_qs.values('responsavel__first_name', 'responsavel__last_name')
            .annotate(total=Count('id'))
            .order_by('-total')[:10]
        )
        
        # Processos recentes
        processos_recentes = processos_qs.select_related(
            'cliente', 'tipo_processo', 'area_direito', 'responsavel'
        ).order_by('-data_inicio')[:20]
        
        context.update({
            'total_processos': total_processos,
            'processos_ativos': processos_ativos,
            'processos_encerrados': processos_encerrados,
            'valor_total_causas': valor_total_causas,
            'processos_por_status': json.dumps(processos_por_status),
            'processos_por_area': processos_por_area,
            'processos_por_responsavel': processos_por_responsavel,
            'processos_recentes': processos_recentes,
        })
        
        return context
    
    def _aplicar_filtros(self, queryset, filtros):
        """
        Aplica filtros ao queryset de processos
        """
        # Filtro por período
        data_inicio, data_fim = self._get_periodo_filtro(filtros)
        if data_inicio:
            queryset = queryset.filter(data_inicio__gte=data_inicio)
        if data_fim:
            queryset = queryset.filter(data_inicio__lte=data_fim)
        
        # Outros filtros
        if filtros.get('tipo_processo'):
            queryset = queryset.filter(tipo_processo=filtros['tipo_processo'])
        
        if filtros.get('area_direito'):
            queryset = queryset.filter(area_direito=filtros['area_direito'])
        
        if filtros.get('status_processo'):
            queryset = queryset.filter(status=filtros['status_processo'])
        
        if filtros.get('responsavel'):
            queryset = queryset.filter(responsavel=filtros['responsavel'])
        
        if filtros.get('valor_causa_min'):
            queryset = queryset.filter(valor_causa__gte=filtros['valor_causa_min'])
        
        if filtros.get('valor_causa_max'):
            queryset = queryset.filter(valor_causa__lte=filtros['valor_causa_max'])
        
        return queryset
    
    def _get_periodo_filtro(self, filtros):
        """
        Converte o período selecionado em datas de início e fim
        """
        periodo = filtros.get('periodo')
        hoje = date.today()
        
        if periodo == 'hoje':
            return hoje, hoje
        elif periodo == 'ontem':
            ontem = hoje - timedelta(days=1)
            return ontem, ontem
        elif periodo == 'esta_semana':
            inicio_semana = hoje - timedelta(days=hoje.weekday())
            return inicio_semana, hoje
        elif periodo == 'semana_passada':
            fim_semana_passada = hoje - timedelta(days=hoje.weekday() + 1)
            inicio_semana_passada = fim_semana_passada - timedelta(days=6)
            return inicio_semana_passada, fim_semana_passada
        elif periodo == 'este_mes':
            inicio_mes = hoje.replace(day=1)
            return inicio_mes, hoje
        elif periodo == 'mes_passado':
            if hoje.month == 1:
                inicio_mes_passado = hoje.replace(year=hoje.year-1, month=12, day=1)
                fim_mes_passado = hoje.replace(day=1) - timedelta(days=1)
            else:
                inicio_mes_passado = hoje.replace(month=hoje.month-1, day=1)
                fim_mes_passado = hoje.replace(day=1) - timedelta(days=1)
            return inicio_mes_passado, fim_mes_passado
        elif periodo == 'ultimo_trimestre':
            # Implementar lógica do trimestre
            return None, None
        elif periodo == 'este_ano':
            inicio_ano = hoje.replace(month=1, day=1)
            return inicio_ano, hoje
        elif periodo == 'ano_passado':
            inicio_ano_passado = hoje.replace(year=hoje.year-1, month=1, day=1)
            fim_ano_passado = hoje.replace(year=hoje.year-1, month=12, day=31)
            return inicio_ano_passado, fim_ano_passado
        elif periodo == 'personalizado':
            return filtros.get('data_inicio'), filtros.get('data_fim')
        
        return None, None


class RelatorioClientesView(LoginRequiredMixin, TemplateView):
    """
    Relatório específico de clientes com análises e estatísticas
    """
    template_name = 'relatorios/clientes.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aplicar filtros
        form = FiltroRelatorioForm(self.request.GET, user=self.request.user, tipo_relatorio='clientes')
        context['form'] = form
        
        # Base queryset
        if self.request.user.is_staff:
            clientes_qs = Cliente.objects.all()
        else:
            # Obter todos os clientes
            clientes_qs = Cliente.objects.all()
        
        if form.is_valid():
            # Aplicar filtros específicos de clientes
            if form.cleaned_data.get('tipo_pessoa'):
                clientes_qs = clientes_qs.filter(tipo_pessoa=form.cleaned_data['tipo_pessoa'])
            
            if form.cleaned_data.get('uf_cliente'):
                clientes_qs = clientes_qs.filter(uf=form.cleaned_data['uf_cliente'])
            
            # Filtro por período de cadastro
            data_inicio, data_fim = self._get_periodo_filtro(form.cleaned_data)
            if data_inicio:
                clientes_qs = clientes_qs.filter(created_at__date__gte=data_inicio)
            if data_fim:
                clientes_qs = clientes_qs.filter(created_at__date__lte=data_fim)
        
        # Estatísticas
        total_clientes = clientes_qs.count()
        clientes_ativos = clientes_qs.filter(ativo=True).count()
        clientes_pf = clientes_qs.filter(tipo_pessoa='PF').count()
        clientes_pj = clientes_qs.filter(tipo_pessoa='PJ').count()
        
        # Clientes por estado
        clientes_por_uf = list(
            clientes_qs.exclude(uf__isnull=True)
            .values('uf')
            .annotate(total=Count('id'))
            .order_by('-total')[:10]
        )
        
        # Clientes com mais processos
        clientes_top = clientes_qs.annotate(
            total_processos=Count('processos')
        ).filter(total_processos__gt=0).order_by('-total_processos')[:10]
        
        # Evolução de cadastros por mês
        cadastros_por_mes = []
        for i in range(12):
            mes = date.today() - timedelta(days=30*i)
            inicio_mes = mes.replace(day=1)
            if mes.month == 12:
                fim_mes = mes.replace(year=mes.year+1, month=1, day=1) - timedelta(days=1)
            else:
                fim_mes = mes.replace(month=mes.month+1, day=1) - timedelta(days=1)
            
            total = clientes_qs.filter(
                created_at__date__gte=inicio_mes,
                created_at__date__lte=fim_mes
            ).count()
            
            cadastros_por_mes.append({
                'mes': mes.strftime('%m/%Y'),
                'total': total
            })
        
        cadastros_por_mes.reverse()
        
        context.update({
            'total_clientes': total_clientes,
            'clientes_ativos': clientes_ativos,
            'clientes_pf': clientes_pf,
            'clientes_pj': clientes_pj,
            'clientes_por_uf': clientes_por_uf,
            'clientes_top': clientes_top,
            'cadastros_por_mes': json.dumps(cadastros_por_mes),
        })
        
        return context
    
    def _get_periodo_filtro(self, filtros):
        # Mesmo método da view de processos
        return RelatorioProcessosView()._get_periodo_filtro(filtros)


class RelatorioFinanceiroView(LoginRequiredMixin, TemplateView):
    """
    Relatório financeiro com gráficos e análises detalhadas
    """
    template_name = 'relatorios/financeiro.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtros
        filtros = {
            'data_inicial': self.request.GET.get('data_inicial'),
            'data_final': self.request.GET.get('data_final'),
            'tipo_relatorio': self.request.GET.get('tipo_relatorio', 'geral'),
        }
        
        # Aplicar período padrão se não especificado
        if not filtros['data_inicial'] or not filtros['data_final']:
            hoje = timezone.now().date()
            filtros['data_inicial'] = hoje.replace(day=1)
            filtros['data_final'] = hoje
        
        # Converter strings para dates
        if isinstance(filtros['data_inicial'], str):
            filtros['data_inicial'] = datetime.strptime(filtros['data_inicial'], '%Y-%m-%d').date()
        if isinstance(filtros['data_final'], str):
            filtros['data_final'] = datetime.strptime(filtros['data_final'], '%Y-%m-%d').date()
        
        # Honorários (receitas)
        honorarios = Honorario.objects.filter(
            usuario=self.request.user,
            data_vencimento__range=[filtros['data_inicial'], filtros['data_final']]
        )
        
        # Despesas
        despesas = Despesa.objects.filter(
            usuario=self.request.user,
            data_despesa__range=[filtros['data_inicial'], filtros['data_final']]
        )
        
        # Cálculos principais
        total_receitas = honorarios.aggregate(total=Sum('valor'))['total'] or Decimal('0')
        total_despesas = despesas.aggregate(total=Sum('valor'))['total'] or Decimal('0')
        lucro_liquido = total_receitas - total_despesas
        margem_lucro = (lucro_liquido / total_receitas * 100) if total_receitas > 0 else 0
        
        # Período anterior para comparação
        periodo_anterior_inicio = filtros['data_inicial'] - timedelta(days=(filtros['data_final'] - filtros['data_inicial']).days + 1)
        periodo_anterior_fim = filtros['data_inicial'] - timedelta(days=1)
        
        honorarios_anterior = Honorario.objects.filter(
            usuario=self.request.user,
            data_vencimento__range=[periodo_anterior_inicio, periodo_anterior_fim]
        )
        despesas_anterior = Despesa.objects.filter(
            usuario=self.request.user,
            data_despesa__range=[periodo_anterior_inicio, periodo_anterior_fim]
        )
        
        total_receitas_anterior = honorarios_anterior.aggregate(total=Sum('valor'))['total'] or Decimal('0')
        total_despesas_anterior = despesas_anterior.aggregate(total=Sum('valor'))['total'] or Decimal('0')
        lucro_anterior = total_receitas_anterior - total_despesas_anterior
        margem_anterior = (lucro_anterior / total_receitas_anterior * 100) if total_receitas_anterior > 0 else 0
        
        # Variações percentuais
        variacao_receitas = ((total_receitas - total_receitas_anterior) / total_receitas_anterior * 100) if total_receitas_anterior > 0 else 0
        variacao_despesas = ((total_despesas - total_despesas_anterior) / total_despesas_anterior * 100) if total_despesas_anterior > 0 else 0
        variacao_lucro = ((lucro_liquido - lucro_anterior) / abs(lucro_anterior) * 100) if lucro_anterior != 0 else 0
        variacao_margem = margem_lucro - margem_anterior
        
        # Dados para gráficos
        dados_evolucao = self._get_dados_evolucao(filtros)
        dados_distribuicao = {
            'receitas': float(total_receitas),
            'despesas': float(total_despesas)
        }
        dados_top_clientes = self._get_top_clientes(filtros)
        dados_despesas_categoria = self._get_despesas_categoria(filtros)
        
        context.update({
            'total_receitas': total_receitas,
            'total_despesas': total_despesas,
            'lucro_liquido': lucro_liquido,
            'margem_lucro': margem_lucro,
            'variacao_receitas': variacao_receitas,
            'variacao_despesas': variacao_despesas,
            'variacao_lucro': variacao_lucro,
            'variacao_margem': variacao_margem,
            'receitas_detalhadas': honorarios[:10],
            'despesas_detalhadas': despesas[:10],
            'dados_evolucao': json.dumps(dados_evolucao),
            'dados_distribuicao': json.dumps(dados_distribuicao),
            'dados_top_clientes': json.dumps(dados_top_clientes),
            'dados_despesas_categoria': json.dumps(dados_despesas_categoria),
        })
        
        return context
    
    def _get_dados_evolucao(self, filtros):
        """Gera dados para o gráfico de evolução financeira"""
        dados = {
            'labels': [],
            'receitas': [],
            'despesas': [],
            'lucro': []
        }
        
        # Gerar dados mensais ou diários baseado no período
        periodo_dias = (filtros['data_final'] - filtros['data_inicial']).days
        
        if periodo_dias <= 31:  # Dados diários
            current_date = filtros['data_inicial']
            while current_date <= filtros['data_final']:
                dados['labels'].append(current_date.strftime('%d/%m'))
                
                receitas_dia = Honorario.objects.filter(
                    usuario=self.request.user,
                    data_vencimento=current_date
                ).aggregate(total=Sum('valor'))['total'] or 0
                
                despesas_dia = Despesa.objects.filter(
                    usuario=self.request.user,
                    data_despesa=current_date
                ).aggregate(total=Sum('valor'))['total'] or 0
                
                dados['receitas'].append(float(receitas_dia))
                dados['despesas'].append(float(despesas_dia))
                dados['lucro'].append(float(receitas_dia - despesas_dia))
                
                current_date += timedelta(days=1)
        else:  # Dados mensais
            current_date = filtros['data_inicial'].replace(day=1)
            while current_date <= filtros['data_final']:
                if current_date.month == 12:
                    next_month = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    next_month = current_date.replace(month=current_date.month + 1)
                
                fim_mes = min(next_month - timedelta(days=1), filtros['data_final'])
                
                dados['labels'].append(current_date.strftime('%m/%Y'))
                
                receitas_mes = Honorario.objects.filter(
                    usuario=self.request.user,
                    data_vencimento__range=[current_date, fim_mes]
                ).aggregate(total=Sum('valor'))['total'] or 0
                
                despesas_mes = Despesa.objects.filter(
                    usuario=self.request.user,
                    data_despesa__range=[current_date, fim_mes]
                ).aggregate(total=Sum('valor'))['total'] or 0
                
                dados['receitas'].append(float(receitas_mes))
                dados['despesas'].append(float(despesas_mes))
                dados['lucro'].append(float(receitas_mes - despesas_mes))
                
                current_date = next_month
        
        return dados
    
    def _get_top_clientes(self, filtros):
        """Gera dados para o gráfico de top clientes"""
        top_clientes = Honorario.objects.filter(
            usuario=self.request.user,
            data_vencimento__range=[filtros['data_inicial'], filtros['data_final']]
        ).values('cliente__nome_razao_social').annotate(
            total=Sum('valor')
        ).order_by('-total')[:5]
        
        return {
            'labels': [cliente['cliente__nome_razao_social'][:20] for cliente in top_clientes],
            'valores': [float(cliente['total']) for cliente in top_clientes]
        }
    
    def _get_despesas_categoria(self, filtros):
        """Gera dados para o gráfico de despesas por categoria"""
        despesas_categoria = Despesa.objects.filter(
            usuario=self.request.user,
            data_despesa__range=[filtros['data_inicial'], filtros['data_final']]
        ).values('tipo_despesa').annotate(
            total=Sum('valor')
        ).order_by('-total')
        
        return {
            'labels': [despesa['tipo_despesa'] for despesa in despesas_categoria],
            'valores': [float(despesa['total']) for despesa in despesas_categoria]
        }
    
    def _get_periodo_filtro(self, filtros):
        # Mesmo método da view de processos
        return RelatorioProcessosView()._get_periodo_filtro(filtros)


class RelatorioProdutividadeView(LoginRequiredMixin, TemplateView):
    """
    Relatório de produtividade com métricas de desempenho
    """
    template_name = 'relatorios/produtividade.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aplicar filtros
        form = FiltroRelatorioForm(self.request.GET, user=self.request.user, tipo_relatorio='produtividade')
        context['form'] = form
        
        # Base querysets
        if self.request.user.is_staff:
            processos_qs = Processo.objects.all()
            andamentos_qs = Andamento.objects.all()
            documentos_qs = Documento.objects.all()
        else:
            processos_qs = Processo.objects.filter(responsavel=self.request.user)
            andamentos_qs = Andamento.objects.filter(processo__responsavel=self.request.user)
            documentos_qs = Documento.objects.filter(processo__responsavel=self.request.user)
        
        if form.is_valid():
            # Aplicar filtros de período
            data_inicio, data_fim = self._get_periodo_filtro(form.cleaned_data)
            if data_inicio:
                processos_qs = processos_qs.filter(data_inicio__gte=data_inicio)
                andamentos_qs = andamentos_qs.filter(data__gte=data_inicio)
                documentos_qs = documentos_qs.filter(data_upload__gte=data_inicio)
            if data_fim:
                processos_qs = processos_qs.filter(data_inicio__lte=data_fim)
                andamentos_qs = andamentos_qs.filter(data__lte=data_fim)
                documentos_qs = documentos_qs.filter(data_upload__lte=data_fim)
        
        # Estatísticas gerais
        total_processos = processos_qs.count()
        total_andamentos = andamentos_qs.count()
        total_documentos = documentos_qs.count()
        
        # Produtividade por usuário (apenas para staff)
        produtividade_usuarios = []
        if self.request.user.is_staff:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            usuarios = User.objects.filter(is_active=True)
            for usuario in usuarios:
                processos_usuario = processos_qs.filter(responsavel=usuario).count()
                andamentos_usuario = andamentos_qs.filter(processo__responsavel=usuario).count()
                documentos_usuario = documentos_qs.filter(processo__responsavel=usuario).count()
                
                if processos_usuario > 0 or andamentos_usuario > 0 or documentos_usuario > 0:
                    produtividade_usuarios.append({
                        'usuario': f"{usuario.first_name} {usuario.last_name}".strip() or usuario.username,
                        'processos': processos_usuario,
                        'andamentos': andamentos_usuario,
                        'documentos': documentos_usuario,
                        'total_atividades': processos_usuario + andamentos_usuario + documentos_usuario
                    })
            
            produtividade_usuarios.sort(key=lambda x: x['total_atividades'], reverse=True)
        
        # Atividades por dia da semana
        atividades_semana = [0] * 7  # Segunda a domingo
        for andamento in andamentos_qs:
            dia_semana = andamento.data.weekday()  # 0=segunda, 6=domingo
            atividades_semana[dia_semana] += 1
        
        # Evolução de atividades por mês
        atividades_por_mes = []
        for i in range(12):
            mes = date.today() - timedelta(days=30*i)
            inicio_mes = mes.replace(day=1)
            if mes.month == 12:
                fim_mes = mes.replace(year=mes.year+1, month=1, day=1) - timedelta(days=1)
            else:
                fim_mes = mes.replace(month=mes.month+1, day=1) - timedelta(days=1)
            
            processos_mes = processos_qs.filter(
                data_inicio__gte=inicio_mes,
                data_inicio__lte=fim_mes
            ).count()
            
            andamentos_mes = andamentos_qs.filter(
                data__gte=inicio_mes,
                data__lte=fim_mes
            ).count()
            
            atividades_por_mes.append({
                'mes': mes.strftime('%m/%Y'),
                'processos': processos_mes,
                'andamentos': andamentos_mes,
                'total': processos_mes + andamentos_mes
            })
        
        atividades_por_mes.reverse()
        
        # Tipos de andamento mais frequentes
        tipos_andamento = list(
            andamentos_qs.values('tipo_andamento__nome')
            .annotate(total=Count('id'))
            .order_by('-total')[:10]
        )
        
        context.update({
            'total_processos': total_processos,
            'total_andamentos': total_andamentos,
            'total_documentos': total_documentos,
            'produtividade_usuarios': produtividade_usuarios[:10],
            'atividades_semana': json.dumps(atividades_semana),
            'atividades_por_mes': json.dumps(atividades_por_mes),
            'tipos_andamento': tipos_andamento,
        })
        
        return context
    
    def _get_periodo_filtro(self, filtros):
        # Mesmo método da view de processos
        return RelatorioProcessosView()._get_periodo_filtro(filtros)


# Views para Templates de Relatórios
class TemplateRelatorioListView(LoginRequiredMixin, ListView):
    """
    Listagem de templates de relatórios
    """
    model = TemplateRelatorio
    template_name = 'relatorios/templates_list.html'
    context_object_name = 'templates'
    paginate_by = 20
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return TemplateRelatorio.objects.all().order_by('-data_criacao')
        else:
            return TemplateRelatorio.objects.filter(
                Q(usuario_criador=self.request.user) | Q(publico=True)
            ).order_by('-data_criacao')


class TemplateRelatorioCreateView(LoginRequiredMixin, CreateView):
    """
    Criação de template de relatório
    """
    model = TemplateRelatorio
    form_class = TemplateRelatorioForm
    template_name = 'relatorios/template_form.html'
    success_url = reverse_lazy('relatorios:templates')
    
    def form_valid(self, form):
        form.instance.usuario_criador = self.request.user
        messages.success(self.request, 'Template de relatório criado com sucesso!')
        return super().form_valid(form)


class TemplateRelatorioUpdateView(LoginRequiredMixin, UpdateView):
    """
    Edição de template de relatório
    """
    model = TemplateRelatorio
    form_class = TemplateRelatorioForm
    template_name = 'relatorios/template_form.html'
    success_url = reverse_lazy('relatorios:templates')
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return TemplateRelatorio.objects.all()
        else:
            return TemplateRelatorio.objects.filter(usuario_criador=self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Template de relatório atualizado com sucesso!')
        return super().form_valid(form)


class TemplateRelatorioDeleteView(LoginRequiredMixin, DeleteView):
    """
    Exclusão de template de relatório
    """
    model = TemplateRelatorio
    template_name = 'relatorios/template_confirm_delete.html'
    success_url = reverse_lazy('relatorios:templates')
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return TemplateRelatorio.objects.all()
        else:
            return TemplateRelatorio.objects.filter(usuario_criador=self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Template de relatório excluído com sucesso!')
        return super().delete(request, *args, **kwargs)


# Views para Dashboards Personalizados
class DashboardPersonalizadoListView(LoginRequiredMixin, ListView):
    """
    Listagem de dashboards personalizados
    """
    model = DashboardPersonalizado
    template_name = 'relatorios/dashboards_list.html'
    context_object_name = 'dashboards'
    paginate_by = 20
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return DashboardPersonalizado.objects.all().order_by('-data_criacao')
        else:
            return DashboardPersonalizado.objects.filter(
                Q(usuario=self.request.user) | Q(publico=True)
            ).order_by('-data_criacao')


class DashboardPersonalizadoCreateView(LoginRequiredMixin, CreateView):
    """
    Criação de dashboard personalizado
    """
    model = DashboardPersonalizado
    form_class = DashboardPersonalizadoForm
    template_name = 'relatorios/dashboard_form.html'
    success_url = reverse_lazy('relatorios:dashboards')
    
    def form_valid(self, form):
        form.instance.usuario = self.request.user
        messages.success(self.request, 'Dashboard personalizado criado com sucesso!')
        return super().form_valid(form)


class DashboardPersonalizadoDetailView(LoginRequiredMixin, DetailView):
    """
    Visualização de dashboard personalizado
    """
    model = DashboardPersonalizado
    template_name = 'relatorios/dashboard_detail.html'
    context_object_name = 'dashboard'
    
    def get_queryset(self):
        if self.request.user.is_staff:
            return DashboardPersonalizado.objects.all()
        else:
            return DashboardPersonalizado.objects.filter(
                Q(usuario=self.request.user) | Q(publico=True)
            )


# Views para Execução de Relatórios
@login_required
def executar_relatorio(request, template_id):
    """
    Executa um relatório baseado em um template
    """
    template = get_object_or_404(TemplateRelatorio, id=template_id)
    
    # Verificar permissões
    if not template.publico and template.usuario_criador != request.user and not request.user.is_staff:
        raise Http404("Template não encontrado")
    
    if request.method == 'POST':
        form = FiltroRelatorioForm(request.POST, user=request.user, tipo_relatorio=template.tipo)
        
        if form.is_valid():
            # Criar execução do relatório
            execucao = ExecucaoRelatorio.objects.create(
                template=template,
                usuario=request.user,
                parametros_execucao=form.cleaned_data,
                status='processando'
            )
            
            try:
                # Gerar dados do relatório
                dados = _gerar_dados_relatorio(template, form.cleaned_data, request.user)
                
                # Atualizar execução
                execucao.status = 'concluido'
                execucao.data_conclusao = timezone.now()
                execucao.total_registros = len(dados.get('registros', []))
                execucao.save()
                
                # Renderizar relatório
                context = {
                    'template': template,
                    'execucao': execucao,
                    'dados': dados,
                    'form': form
                }
                
                return render(request, 'relatorios/resultado.html', context)
                
            except Exception as e:
                execucao.status = 'erro'
                execucao.mensagem_erro = str(e)
                execucao.data_conclusao = timezone.now()
                execucao.save()
                
                messages.error(request, f'Erro ao executar relatório: {str(e)}')
                return redirect('relatorios:dashboard')
    else:
        form = FiltroRelatorioForm(user=request.user, tipo_relatorio=template.tipo)
    
    context = {
        'template': template,
        'form': form
    }
    
    return render(request, 'relatorios/executar.html', context)


def _gerar_dados_relatorio(template, filtros, usuario):
    """
    Gera os dados do relatório baseado no template e filtros
    """
    dados = {
        'registros': [],
        'estatisticas': {},
        'graficos': []
    }
    
    # Aplicar filtros de período
    data_inicio, data_fim = _get_periodo_filtro_helper(filtros)
    
    if template.tipo == 'processos':
        # Base queryset
        if usuario.is_staff:
            qs = Processo.objects.all()
        else:
            qs = Processo.objects.filter(responsavel=usuario)
        
        # Aplicar filtros
        if data_inicio:
            qs = qs.filter(data_inicio__gte=data_inicio)
        if data_fim:
            qs = qs.filter(data_inicio__lte=data_fim)
        
        # Aplicar outros filtros específicos
        if filtros.get('tipo_processo'):
            qs = qs.filter(tipo_processo=filtros['tipo_processo'])
        if filtros.get('area_direito'):
            qs = qs.filter(area_direito=filtros['area_direito'])
        if filtros.get('status_processo'):
            qs = qs.filter(status=filtros['status_processo'])
        
        # Selecionar campos baseado no template
        campos_selecionados = template.campos_selecionados or []
        
        # Gerar registros
        for processo in qs.select_related('cliente', 'area_direito', 'responsavel'):
            registro = {}
            
            if 'numero_processo' in campos_selecionados:
                registro['numero_processo'] = processo.numero_processo
            if 'cliente' in campos_selecionados:
                registro['cliente'] = str(processo.cliente)
            if 'tipo_processo' in campos_selecionados:
                registro['tipo_processo'] = str(processo.tipo_processo)
            if 'area_direito' in campos_selecionados:
                registro['area_direito'] = str(processo.area_direito)
            if 'status' in campos_selecionados:
                registro['status'] = processo.get_status_display()
            if 'data_inicio' in campos_selecionados:
                registro['data_inicio'] = processo.data_inicio
            if 'valor_causa' in campos_selecionados:
                registro['valor_causa'] = processo.valor_causa
            if 'responsavel' in campos_selecionados:
                registro['responsavel'] = str(processo.responsavel)
            
            dados['registros'].append(registro)
        
        # Estatísticas
        dados['estatisticas'] = {
            'total_processos': qs.count(),
            'processos_ativos': qs.filter(status='ativo').count(),
            'processos_encerrados': qs.filter(status='encerrado').count(),
            'valor_total_causas': qs.aggregate(Sum('valor_causa'))['valor_causa__sum'] or Decimal('0.00')
        }
    
    elif template.tipo == 'clientes':
        # Implementar lógica para relatório de clientes
        # Obter todos os clientes
        qs = Cliente.objects.all()
        
        # Aplicar filtros de período
        if data_inicio:
            qs = qs.filter(created_at__date__gte=data_inicio)
        if data_fim:
            qs = qs.filter(created_at__date__lte=data_fim)
        
        # Gerar registros
        campos_selecionados = template.campos_selecionados or []
        
        for cliente in qs:
            registro = {}
            
            if 'nome' in campos_selecionados:
                registro['nome'] = cliente.nome
            if 'tipo_pessoa' in campos_selecionados:
                registro['tipo_pessoa'] = cliente.get_tipo_pessoa_display()
            if 'documento' in campos_selecionados:
                registro['documento'] = cliente.cpf_cnpj
            if 'email' in campos_selecionados:
                registro['email'] = cliente.email
            if 'telefone' in campos_selecionados:
                registro['telefone'] = cliente.telefone
            if 'data_cadastro' in campos_selecionados:
                registro['data_cadastro'] = cliente.created_at
            
            dados['registros'].append(registro)
        
        # Estatísticas
        dados['estatisticas'] = {
            'total_clientes': qs.count(),
            'clientes_pf': qs.filter(tipo_pessoa='PF').count(),
            'clientes_pj': qs.filter(tipo_pessoa='PJ').count(),
            'clientes_ativos': qs.filter(ativo=True).count()
        }
    
    elif template.tipo == 'financeiro':
        # Implementar lógica para relatório financeiro
        # Obter todos os honorários e despesas
        honorarios_qs = Honorario.objects.all()
        despesas_qs = Despesa.objects.all()
        
        # Aplicar filtros de período
        if data_inicio:
            honorarios_qs = honorarios_qs.filter(data_vencimento__gte=data_inicio)
            despesas_qs = despesas_qs.filter(data_vencimento__gte=data_inicio)
        if data_fim:
            honorarios_qs = honorarios_qs.filter(data_vencimento__lte=data_fim)
            despesas_qs = despesas_qs.filter(data_vencimento__lte=data_fim)
        
        # Combinar honorários e despesas
        campos_selecionados = template.campos_selecionados or []
        
        # Honorários
        for honorario in honorarios_qs.select_related('processo', 'processo__cliente'):
            registro = {'tipo': 'Honorário'}
            
            if 'processo' in campos_selecionados:
                registro['processo'] = honorario.processo.numero_processo
            if 'cliente' in campos_selecionados:
                registro['cliente'] = str(honorario.processo.cliente)
            if 'valor' in campos_selecionados:
                registro['valor'] = honorario.valor_total
            if 'data_vencimento' in campos_selecionados:
                registro['data_vencimento'] = honorario.data_vencimento
            if 'status' in campos_selecionados:
                registro['status'] = honorario.get_status_display()
            
            dados['registros'].append(registro)
        
        # Despesas
        for despesa in despesas_qs.select_related('processo', 'processo__cliente', 'tipo_despesa'):
            registro = {'tipo': 'Despesa'}
            
            if 'processo' in campos_selecionados:
                registro['processo'] = despesa.processo.numero_processo
            if 'cliente' in campos_selecionados:
                registro['cliente'] = str(despesa.processo.cliente)
            if 'valor' in campos_selecionados:
                registro['valor'] = despesa.valor
            if 'data_vencimento' in campos_selecionados:
                registro['data_vencimento'] = despesa.data_vencimento
            if 'descricao' in campos_selecionados:
                registro['descricao'] = despesa.descricao
            
            dados['registros'].append(registro)
        
        # Estatísticas
        total_honorarios = honorarios_qs.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0.00')
        total_despesas = despesas_qs.aggregate(Sum('valor'))['valor__sum'] or Decimal('0.00')
        
        dados['estatisticas'] = {
            'total_honorarios': total_honorarios,
            'total_despesas': total_despesas,
            'saldo': total_honorarios - total_despesas,
            'honorarios_pendentes': honorarios_qs.filter(status='pendente').count(),
            'despesas_pendentes': despesas_qs.filter(reembolsada=False).count()
        }
    
    return dados


def _get_periodo_filtro_helper(filtros):
    """
    Helper para converter período em datas
    """
    return RelatorioProcessosView()._get_periodo_filtro(filtros)


# Views de Exportação
@login_required
def exportar_relatorio(request, execucao_id):
    """
    Exporta um relatório executado em diferentes formatos
    """
    execucao = get_object_or_404(ExecucaoRelatorio, id=execucao_id, usuario=request.user)
    
    if request.method == 'POST':
        form = ExportarRelatorioForm(request.POST)
        
        if form.is_valid():
            formato = form.cleaned_data['formato']
            
            # Regenerar dados do relatório
            dados = _gerar_dados_relatorio(
                execucao.template,
                execucao.parametros_execucao,
                request.user
            )
            
            if formato == 'pdf':
                return _exportar_pdf(execucao, dados, form.cleaned_data)
            elif formato == 'excel':
                return _exportar_excel(execucao, dados, form.cleaned_data)
            elif formato == 'csv':
                return _exportar_csv(execucao, dados, form.cleaned_data)
            else:
                messages.error(request, 'Formato de exportação inválido.')
    else:
        form = ExportarRelatorioForm()
    
    context = {
        'execucao': execucao,
        'form': form
    }
    
    return render(request, 'relatorios/exportar.html', context)


def _exportar_pdf(execucao, dados, opcoes):
    """
    Exporta relatório em formato PDF
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{execucao.id}.pdf"'
    
    # Criar documento PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Título
    titulo = opcoes.get('titulo_personalizado') or f"Relatório: {execucao.template.nome}"
    story.append(Paragraph(titulo, styles['Title']))
    story.append(Spacer(1, 12))
    
    # Informações do relatório
    info_data = [
        ['Template:', execucao.template.nome],
        ['Executado em:', execucao.data_execucao.strftime('%d/%m/%Y %H:%M')],
        ['Total de registros:', str(execucao.total_registros)],
    ]
    
    info_table = Table(info_data)
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 12))
    
    # Dados do relatório
    if dados['registros']:
        # Cabeçalhos da tabela
        headers = list(dados['registros'][0].keys())
        table_data = [headers]
        
        # Dados
        for registro in dados['registros'][:100]:  # Limitar a 100 registros
            row = [str(registro.get(header, '')) for header in headers]
            table_data.append(row)
        
        # Criar tabela
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
    
    # Construir PDF
    doc.build(story)
    
    return response


def _exportar_excel(execucao, dados, opcoes):
    """
    Exporta relatório em formato Excel
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="relatorio_{execucao.id}.xlsx"'
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Título
    titulo = opcoes.get('titulo_personalizado') or f"Relatório: {execucao.template.nome}"
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    
    # Informações do relatório
    ws['A3'] = "Template:"
    ws['B3'] = execucao.template.nome
    ws['A4'] = "Executado em:"
    ws['B4'] = execucao.data_execucao.strftime('%d/%m/%Y %H:%M')
    ws['A5'] = "Total de registros:"
    ws['B5'] = execucao.total_registros
    
    # Dados do relatório
    if dados['registros']:
        start_row = 7
        
        # Cabeçalhos
        headers = list(dados['registros'][0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Dados
        for row_idx, registro in enumerate(dados['registros'], start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(registro.get(header, '')))
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar workbook
    wb.save(response)
    
    return response


def _exportar_csv(execucao, dados, opcoes):
    """
    Exporta relatório em formato CSV
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{execucao.id}.csv"'
    
    writer = csv.writer(response)
    
    # Cabeçalho com informações do relatório
    writer.writerow([f"Relatório: {execucao.template.nome}"])
    writer.writerow([f"Executado em: {execucao.data_execucao.strftime('%d/%m/%Y %H:%M')}"])
    writer.writerow([f"Total de registros: {execucao.total_registros}"])
    writer.writerow([])  # Linha em branco
    
    # Dados do relatório
    if dados['registros']:
        # Cabeçalhos
        headers = list(dados['registros'][0].keys())
        writer.writerow(headers)
        
        # Dados
        for registro in dados['registros']:
            row = [str(registro.get(header, '')) for header in headers]
            writer.writerow(row)
    
    return response
